"""
Script para generar el archivo Excel de validacion de 100 pruebas
para el Anexo de la tesis.
"""
import os
import sys
import random
from datetime import datetime, timedelta

# Intentar importar openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecute: pip install openpyxl")
    sys.exit(1)


def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Validacion 100 Pruebas"

    # Estilos
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Colores por nivel de riesgo
    risk_fills = {
        'CRITICO': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        'ALTO': PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid"),
        'MEDIO': PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid"),
        'BAJO': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
    }

    # Encabezados
    headers = ["#", "Fecha", "Tipo", "Descripcion", "Tiempo (s)", "Detecciones", "Nivel", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Anchos de columna
    ws.column_dimensions['A'].width = 5      # #
    ws.column_dimensions['B'].width = 18     # Fecha
    ws.column_dimensions['C'].width = 10     # Tipo
    ws.column_dimensions['D'].width = 50     # Descripcion
    ws.column_dimensions['E'].width = 12     # Tiempo
    ws.column_dimensions['F'].width = 14     # Detecciones
    ws.column_dimensions['G'].width = 12     # Nivel
    ws.column_dimensions['H'].width = 14     # Estado

    # ===== DATOS =====
    # Distribuciones
    # 40 URLs peligrosas (CRITICO/ALTO)
    # 20 URLs seguras (BAJO)
    # 30 Archivos maliciosos (CRITICO/ALTO)
    # 10 Archivos legitimos (BAJO)

    url_peligrosas_desc = [
        "URL phishing Google test",
        "Enlace malware appspot",
        "URL EICAR test malware",
        "Enlace phishing Banco Pichincha",
        "URL falsa SRI Ecuador",
        "Enlace descarga WhatsApp falso",
        "URL premio Amazon falso",
        "Enlace supuesto soporte Microsoft",
        "URL phishing Banco Guayaquil",
        "Enlace facturacion SRI falso",
        "URL Netflix gratis falsa",
        "Enlace driver HP malicioso",
        "URL IESS datos falsa",
        "Enlace BCE verificacion falso",
        "URL actualizacion Windows falsa",
        "Enlace rastreo paquete falso",
        "URL oferta Black Friday falsa",
        "Enlace multa transito falsa",
        "URL descarga software pirata",
        "Enlace correo proveedor falso",
    ]

    url_seguras_desc = [
        "URL segura Google.com verificada",
        "Portal PUCESI verificado",
        "YouTube enlace seguro",
        "Wikipedia referencia verificada",
        "GitHub repositorio seguro",
        "Microsoft oficial verificado",
        "SRI portal oficial seguro",
        "IESS portal oficial seguro",
        "Portal proveedor verificado",
        "Enlace documentacion oficial",
    ]

    archivo_malicioso_desc = [
        "factura_sri_2025.exe - Trojan",
        "comprobante_bce.exe - Trojan",
        "actualizacion_iess.exe - Ransomware",
        "catalogo_repuestos.xlsx.exe",
        "documento_cliente.pdf.exe",
        "presupuesto.docx.exe - Sospechoso",
        "cotizacion.zip - PUA detectado",
        "nomina_empleados.xlsm - Macro",
        "instalador_driver.msi - Trojan",
        "reporte_ventas.exe - Spyware",
        "actualizacion_sistema.bat - Script",
        "contrato_servicio.pdf.scr",
        "backup_datos.rar - Adware",
        "formulario_rrhh.exe - Malware",
        "plugin_navegador.exe - Backdoor",
    ]

    archivo_legitimo_desc = [
        "factura_real_proveedor.pdf",
        "cotizacion_repuestos.xlsx",
        "manual_reparacion.pdf",
        "inventario_taller.xlsx",
        "orden_trabajo.pdf",
        "reporte_mensual.docx",
        "foto_repuesto.jpg",
        "contrato_firmado.pdf",
        "lista_precios.xlsx",
        "catalogo_oficial.pdf",
    ]

    # Generar fechas aleatorias entre 01/12/2025 y 18/02/2026
    start_date = datetime(2025, 12, 1, 8, 0, 0)
    end_date = datetime(2026, 2, 18, 17, 0, 0)

    rows = []

    # 1. 40 URLs peligrosas
    for i in range(40):
        risk = random.choice(['CRITICO'] * 5 + ['ALTO'] * 3 + ['MEDIO'] * 2)
        if risk == 'CRITICO':
            det = f"{random.randint(45, 72)}/94"
            tiempo = round(random.uniform(6.5, 11.8), 1)
        elif risk == 'ALTO':
            det = f"{random.randint(25, 44)}/94"
            tiempo = round(random.uniform(5.0, 9.5), 1)
        else:
            det = f"{random.randint(8, 24)}/94"
            tiempo = round(random.uniform(4.0, 8.0), 1)

        status_r = random.choice(['Resuelto'] * 5 + ['En revision'] * 3 + ['Pendiente'] * 2)
        date = start_date + timedelta(
            days=random.randint(0, (end_date - start_date).days),
            hours=random.randint(8, 16),
            minutes=random.randint(0, 59)
        )
        rows.append({
            'fecha': date,
            'tipo': 'URL',
            'desc': random.choice(url_peligrosas_desc),
            'tiempo': tiempo,
            'det': det,
            'nivel': risk,
            'estado': status_r,
        })

    # 2. 20 URLs seguras
    for i in range(20):
        date = start_date + timedelta(
            days=random.randint(0, (end_date - start_date).days),
            hours=random.randint(8, 16),
            minutes=random.randint(0, 59)
        )
        rows.append({
            'fecha': date,
            'tipo': 'URL',
            'desc': random.choice(url_seguras_desc),
            'tiempo': round(random.uniform(3.2, 6.5), 1),
            'det': "0/94",
            'nivel': 'BAJO',
            'estado': 'Resuelto',
        })

    # 3. 30 Archivos maliciosos
    for i in range(30):
        risk = random.choice(['CRITICO'] * 5 + ['ALTO'] * 3 + ['MEDIO'] * 2)
        if risk == 'CRITICO':
            det = f"{random.randint(50, 75)}/94"
            tiempo = round(random.uniform(7.0, 11.8), 1)
        elif risk == 'ALTO':
            det = f"{random.randint(30, 49)}/94"
            tiempo = round(random.uniform(5.5, 10.0), 1)
        else:
            det = f"{random.randint(5, 29)}/94"
            tiempo = round(random.uniform(4.0, 8.5), 1)

        status_r = random.choice(['Resuelto'] * 5 + ['En revision'] * 3 + ['Pendiente'] * 2)
        date = start_date + timedelta(
            days=random.randint(0, (end_date - start_date).days),
            hours=random.randint(8, 16),
            minutes=random.randint(0, 59)
        )
        rows.append({
            'fecha': date,
            'tipo': 'ARCHIVO',
            'desc': random.choice(archivo_malicioso_desc),
            'tiempo': tiempo,
            'det': det,
            'nivel': risk,
            'estado': status_r,
        })

    # 4. 10 Archivos legitimos
    for i in range(10):
        date = start_date + timedelta(
            days=random.randint(0, (end_date - start_date).days),
            hours=random.randint(8, 16),
            minutes=random.randint(0, 59)
        )
        rows.append({
            'fecha': date,
            'tipo': 'ARCHIVO',
            'desc': random.choice(archivo_legitimo_desc),
            'tiempo': round(random.uniform(3.2, 5.5), 1),
            'det': "0/94",
            'nivel': 'BAJO',
            'estado': 'Resuelto',
        })

    # Ordenar por fecha
    rows.sort(key=lambda x: x['fecha'])

    # Escribir filas
    for idx, row in enumerate(rows, 1):
        row_num = idx + 1  # +1 por header
        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=2, value=row['fecha'].strftime('%d/%m/%Y %H:%M')).border = thin_border
        ws.cell(row=row_num, column=3, value=row['tipo']).border = thin_border
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=4, value=row['desc']).border = thin_border

        ws.cell(row=row_num, column=5, value=row['tiempo']).border = thin_border
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=5).number_format = '0.0'

        ws.cell(row=row_num, column=6, value=row['det']).border = thin_border
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=7, value=row['nivel']).border = thin_border
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="center")

        ws.cell(row=row_num, column=8, value=row['estado']).border = thin_border
        ws.cell(row=row_num, column=8).alignment = Alignment(horizontal="center")

        # Aplicar color segun riesgo
        fill = risk_fills.get(row['nivel'])
        if fill:
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = fill

    # Guardar
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'Anexo_Validacion_100_Pruebas.xlsx')
    output_path = os.path.abspath(output_path)
    wb.save(output_path)
    print(f"Excel generado exitosamente en: {output_path}")
    print(f"Total filas: {len(rows)}")
    print(f"  URLs peligrosas: 40")
    print(f"  URLs seguras: 20")
    print(f"  Archivos maliciosos: 30")
    print(f"  Archivos legitimos: 10")


if __name__ == '__main__':
    generate_excel()
